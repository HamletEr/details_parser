import re
import os
import openpyxl

import fpdf


def convert_xlsx_to_pdf(input_file_path: str) -> None:
    # using fpdf2
    # https://pypi.org/project/fpdf2/
    # docs: https://py-pdf.github.io/fpdf2/Tutorial-ru.html

    # создаем pdf-файл
    pdf = fpdf.FPDF(orientation="P", unit="mm", format="A4")

    # добавляем шрифт Roboto, который адекватно воспринимает кириллицу
    pdf.add_font("Roboto", style="", fname="resurses/Fonts/Roboto/Roboto-Regular.ttf", uni=True)
    pdf.add_font("Roboto", style="B", fname="resurses/Fonts/Roboto/Roboto-Bold.ttf", uni=True)
    pdf.add_font("Roboto", style="I", fname="resurses/Fonts/Roboto/Roboto-Italic.ttf", uni=True)
    pdf.add_font("Roboto", style="BI", fname="resurses/Fonts/Roboto/Roboto-BoldItalic.ttf", uni=True)

    pdf.set_font("Roboto", size=6)
    pdf.core_fonts_encoding = 'utf-8'

    # Создаем лист
    pdf.add_page()

    # Задаем стиль таблицы
    pdf.set_draw_color(0, 0, 0)         # цвет границы
    pdf.set_line_width(0.3)

    # Работаем с Excel-файлом ------------------------------------------------------------

    # Загружаем исходный xlsx-файл. Data_only - т.к. формулы нас не интересуют
    input_wb = openpyxl.load_workbook(filename=input_file_path, data_only=True)
    # выбираем первую таблицу
    input_page_1 = input_wb["Table 1"]

    with pdf.table() as pdf_table:                          # Создаем таблицу в pdf

        for xlsx_row in input_page_1:                       # Проходим по строкам в excel
            pdf_table_row = pdf_table.row()                 # Создаем строку в таблице в pdf
            last_pdf_cell_span = 1                      # Размер последней добавленной в pdf ячейки по горизонтали
            for xlsx_cell in xlsx_row:                      # Проходим по ячейкам в строке excel
                print(xlsx_cell.value, end=' ')

                if xlsx_cell.value is not None:
                    last_pdf_cell = pdf_table_row.cell(str(xlsx_cell.value))
                    #                                       Создаем ячейку в pdf с данными из ячейки excel
                else:
                    # TODO здесь мы должны обединить ячейку с предидущей - convert_rowspan ?
                    pass
            print()

    # with pdf.table() as table:
    #     for data_row in data:
    #         row = table.row()
    #         for datum in data_row:
    #             if datum != 'None':
    #                 row.cell(datum)
    #             else:
    #                 row.cell('')

    # сохраняем pdf-файл
    pdf.output(name=f'{input_file_path.replace(".xlsx", ".pdf")}')


def cap_handler(element_data: tuple, output_dir_path: str) -> None:
    """
    Обработчик для Заглушек (CAP)
    """

    from utils.utils import generate_detail_file_name

    template_path = 'templates/CAP 12 SCH 30 BD ASTM A420 GR.WPL6.xlsx'

    template_wb = openpyxl.load_workbook(template_path)

    # Заполняем первую таблицу
    table_1 = template_wb['Table 1']

    table_1['E3'] = re.search(r'CAP.+', element_data[1]).string
    # TODO Здесь есть проблема с переносом строки внутри ячейки.
    #  Не все данные могут быть видны в excel без наведения на ячейку
    table_1['E4'] = element_data[4]
    table_1['E5'] = element_data[2]

    # Заполняем вторую таблицу
    table_2 = template_wb['Table 2']

    # TODO Дописать порядок заполнения обоих таблиц

    # генерируем имя файла
    xlsx_file_name = f'{generate_detail_file_name(element_data)}.xlsx'

    # полный путь для сохранения файла:
    output_xlsx_file_path = os.path.join(output_dir_path, xlsx_file_name)

    # сохраняем excel-файл
    template_wb.save(filename=output_xlsx_file_path)
    template_wb.close()

    # конвертируем сохраненный excel-файл в pdf
    # convert_xlsx_to_pdf(output_xlsx_file_path)
