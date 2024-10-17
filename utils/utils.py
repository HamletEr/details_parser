import os
import re
import openpyxl

from openpyxl.styles import Font

from .specialized_handlers import cap_handler


def get_output_directory(input_file_path: str) -> str:
    """
    Функция создает новую директорию "output" в директории с исходным файлом, если ее еще не существует.
    И возвращает путь к этой директории.

    :param input_file_path: Путь к исходному файлу
    :return: Путь к директории для сохранения новых файлов
    """

    # определяем директорию исходного файла
    input_file_dir = os.path.dirname(input_file_path)

    # создаем рядом с исходным файлом директорию 'output' если ее еще не существует
    os.makedirs(os.path.join(input_file_dir, 'output'), exist_ok=True)

    # получаем путь к директории 'output'
    output_directory = os.path.join(input_file_dir, 'output')

    return output_directory


def copy_excel_with_line_numbers(input_file_path: str, output_dir_path: str, start_row: int = 3) -> str:
    """
    Функция создает копию исходного EXCEL-документа в папке "output", добавляя столбец с номерами строк.
    Нумерация начинается со строки start_row.

    :param input_file_path: Путь к исходному файлу.
    :param output_dir_path: Директория для сохранения выходных файлов (рядом со входным файлом).
    :param start_row: Ряд с которого нужно начать нумерацию.
    :return: Возвращает путь к новому Excel-файлу.
    """
    wb = openpyxl.load_workbook(input_file_path)
    ws = wb.active

    # Добавляем столбец с индексом значимых строк
    ws.insert_cols(0)

    # Корректируем ширину столбцов
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 5

    # Копируем стиль шрифта из имеющейся ячейки таблицы в ячейки нового столбца
    cell_style = ws.cell(row=3, column=3).font  # Ячейка из которой берется образец шрифта

    for i_row, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        row[0].value = i_row + 1
        row[0].font = Font(name=cell_style.name, size=int(cell_style.sz))

    # Указываем имя нового файла
    new_filename = (os.path.basename(input_file_path)
                    .replace('.xlsx', '_indexed.xlsx'))

    # Сохраняем новый проиндексированный файл в директории 'output' рядом с исходным файлом.
    new_file_path = os.path.join(output_dir_path, new_filename)

    wb.save(new_file_path)
    wb.close()

    print(f'Создана копия исходного файла: {new_file_path}')

    return new_file_path


def generate_detail_file_name(element_data: tuple) -> str:
    """
    Функция генерирует корректное имя файла из исходной строки с данными детали
    """

    default_element_name = element_data[1]

    symbols_to_replace = r'[^\w]'       # любые символы, кроме цифр, букв и знака '_'

    correct_name = (f'{element_data[0]}_'
                    f'{re.sub(pattern=symbols_to_replace, repl="_", string=default_element_name)}')

    return correct_name


def main_handler(element_data: tuple, output_dir_path: str) -> None:
    """
    Основной обработчик строк с данными детали.
    Генерирует новые excel и pdf файлы с подробной информацией о детали.

    :param element_data: Данные о детали.
    :param output_dir_path: Директория для сохранения итогового файла.
    """

    if element_data[1] and 'CAP' in element_data[1]:
        cap_handler(element_data, output_dir_path)

    else:
        print(f'Не определен тип детали для строки {element_data[0]}: {element_data}')
