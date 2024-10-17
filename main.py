import openpyxl

from utils.utils import copy_excel_with_line_numbers, main_handler, get_output_directory


def main(input_file_path: str, first_relevant_line: int = 3, need_indexing: bool = False, pause: int = 0) -> None:
    """
    Функция запускает парсинг активного листа Excel-файла "input_file_path".
    Файл обрабатывается начиная со строки start_line.
    Если в файле нет нумерации строк - следует передать параметр need_indexing = True. В этом случае будет
    создана копия исходного файла с нумерацией значимых строк.
    Для каждой значащей строки создает новый excel-файл с подробным описанием изделия,
    а так же аналогичный pdf-файл.
    Если передан параметр pause (N) - функция делает паузу после обработки каждой N-ной строки.
    Все новые файлы сохраняются в директории "output", которая создается автоматически в директории с исходным файлом.

    :param input_file_path: Путь к исходному файлу
    :param first_relevant_line: Первая значимая строка (не заголовок и не пустая строка)
    :param need_indexing: Требуется ли нумерация строк.
    :param pause: Частота паузы между обработкой строк.
    """

    # Получаем директорию "output" для сохранения созданных файлов
    output_dir_path = get_output_directory(input_file_path)

    # Если нужна индексация (нумерация строк) создаем новый файл с индексами значащих строк и далее работаем с ним.
    if need_indexing:
        input_file_path = copy_excel_with_line_numbers(input_file_path=input_file_path,
                                                       output_dir_path=output_dir_path,
                                                       start_row=first_relevant_line)

    # Открываем excel файл и выбираем активный лист со списком деталей
    wb = openpyxl.load_workbook(input_file_path)    # workbook
    ws = wb.active                                  # worksheet

    for element_data in ws.iter_rows(min_row=first_relevant_line, max_row=ws.max_row, values_only=True):

        # передаем данные из текущей строки на дальнейшую обработку
        main_handler(element_data, output_dir_path)

        # пауза (опция)
        if pause and pause > 0:
            if element_data[0] % pause == 0:

                continue_answer = ''
                while continue_answer.lower() not in ('y', 'n'):
                    continue_answer = str(input(f'Обработано {element_data[0]} из '
                                                f'{ws.max_row - first_relevant_line + 1} строк. '
                                                f'Продолжить? (y/n)\n'))

                if continue_answer.lower() == 'n':
                    break
    wb.close()


if __name__ == '__main__':
    main(input_file_path='Inq_2432_CS.xlsx',
         first_relevant_line=3,
         need_indexing=True,
         pause=1)
